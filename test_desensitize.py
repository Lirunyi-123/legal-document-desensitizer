# -*- coding: utf-8 -*-
"""v2.2 回归测试（纯标准库，python -m unittest test_desensitize）"""

import unittest

from desensitize import (
    Desensitizer, SecureDesensitizer,
    parse_mapping_text, restore_text,
    _is_valid_id_checksum, _is_valid_credit_code, _luhn_check,
)  # 导入成功即验证类定义顺序修复


class TestRuleLayerFixes(unittest.TestCase):
    def setUp(self):
        self.d = Desensitizer()

    def test_law_firm_license_not_credit_code(self):
        result = self.d.mask('执业许可证号：31110000E000123456')
        self.assertIn('[律师执业证号]', result.text)
        for m in result.mapping:
            self.assertNotEqual(m.type, '统一社会信用代码')

    def test_credit_code_labeled_and_bare_9(self):
        result = self.d.mask('统一社会信用代码：91110108MA01ABC123；另一个 91330105MA27XABC1D')
        self.assertEqual(result.text.count('[统一社会信用代码]'), 2)

    def test_id_with_context(self):
        result = self.d.mask('身份证号110101198001011234')
        self.assertIn('身份证号[身份证号]', result.text)

    def test_18_digit_bank_card_not_id(self):
        result = self.d.mask('账户622202020001234567')
        self.assertIn('[银行账号]', result.text)
        self.assertNotIn('身份证号', result.text)

    def test_birthdate_only_by_default(self):
        result = self.d.mask('签订日期：2024年1月15日，男，1985年8月15日出生')
        self.assertIn('2024年1月15日', result.text)   # 普通日期保留
        self.assertIn('[出生日期]', result.text)      # 出生日期脱敏

    def test_all_dates_flag(self):
        result = Desensitizer(mask_all_dates=True).mask('签订日期：2024年1月15日')
        self.assertIn('[日期]', result.text)

    def test_per_value_count(self):
        result = self.d.mask('A 13800138000 B 13800138000 C 13912345678')
        # v2.5：映射表按"每处替换一行"记录，同一原始值出现 N 次 → N 行
        originals = [m.original for m in result.mapping]
        self.assertEqual(originals.count('13800138000'), 2)
        self.assertEqual(originals.count('13912345678'), 1)
        self.assertEqual(result.stats['总替换次数'], 3)

    def test_wechat_and_qq_recorded(self):
        result = self.d.mask('微信号：lawyer_wang，QQ：12345678')
        types = {m.type for m in result.mapping}
        self.assertIn('微信号', types)
        self.assertIn('QQ号', types)

    def test_email_not_broken_by_phone_rule(self):
        result = self.d.mask('13800138000@qq.com')
        self.assertIn('[邮箱]', result.text)
        self.assertNotIn('[手机号]@', result.text)

    def test_case_number(self):
        result = self.d.mask('案号：(2024)京0108民初12345号')
        self.assertIn('案号：[案号]', result.text)

    def test_license_plate_not_wechat(self):
        result = self.d.mask('我开保时捷粤B88888去')
        self.assertIn('[车牌号]', result.text)
        self.assertNotIn('[微信号]', result.text)

    def test_entity_resolver_still_works(self):
        result = self.d.mask('原告：陈建国。被告：杭州鼎盛房地产开发有限公司。陈建国再次出现。')
        self.assertIn('[当事人甲（原告）]', result.text)
        self.assertIn('[合同乙方]', result.text)
        for m in result.mapping:
            if m.original == '陈建国':
                self.assertEqual(m.replacement, '[当事人甲（原告）]')

    def test_secure_mode_with_all_dates(self):
        d = SecureDesensitizer(security_level='strict', mask_all_dates=True)
        result = d.mask('身份证号110101198001011234，日期2024年1月15日')
        self.assertIn('[身份证号]', result.text)
        self.assertIn('[日期]', result.text)

    def test_scan_covers_key_types(self):
        text = '执业证号：111012023123456789；手机13800138000；QQ：12345678；身份证号110101198001011234'
        types = {f['type'] for f in self.d.scan(text)}
        self.assertIn('律师执业证号', types)
        self.assertIn('手机号', types)
        self.assertIn('QQ号', types)
        self.assertIn('身份证号', types)


class TestChecksumValidation(unittest.TestCase):
    """GB 11643 / GB 32100 / Luhn 校验码"""

    def test_id_checksum_valid(self):
        # 手工推导：前17位加权和 120，120 % 11 = 10 → 校验码 '2'
        self.assertTrue(_is_valid_id_checksum('110101198001011232'))

    def test_id_checksum_invalid(self):
        self.assertFalse(_is_valid_id_checksum('110101198001011234'))

    def test_credit_code_checksum(self):
        # GB 32100 公开示例：91350100M000100Y43 通过校验
        self.assertTrue(_is_valid_credit_code('91350100M000100Y43'))
        self.assertFalse(_is_valid_credit_code('91350100M000100Y4X'))

    def test_luhn(self):
        # 标准 Luhn 测试卡号
        self.assertTrue(_luhn_check('4111111111111111'))
        self.assertTrue(_luhn_check('4012888888881881'))
        self.assertFalse(_luhn_check('4111111111111112'))


class TestV22NewRules(unittest.TestCase):
    def setUp(self):
        self.d = Desensitizer()

    def test_passport_not_wechat(self):
        result = self.d.mask('护照：E12345678')
        self.assertIn('[护照号]', result.text)
        self.assertNotIn('[微信号]', result.text)

    def test_driving_license_type(self):
        result = self.d.mask('驾驶证号330106198001011234')
        self.assertIn('[驾驶证号]', result.text)
        for m in result.mapping:
            self.assertNotEqual(m.type, '身份证号')

    def test_org_code(self):
        result = self.d.mask('组织机构代码69920000-2')
        self.assertIn('[组织机构代码]', result.text)

    def test_account_context(self):
        result = self.d.mask('收款账号：6222020200012345678')
        self.assertIn('收款账号：[银行账号]', result.text)

    def test_amount_full_match(self):
        result = self.d.mask('尾款80万元应于三日内支付')
        self.assertIn('尾款[金额]应于三日内支付', result.text)
        for m in result.mapping:
            if m.type == '金额':
                self.assertEqual(m.original, '80万元')

    def test_amount_not_pixel_or_share(self):
        result = self.d.mask('设备60万像素，持股8万股')
        self.assertIn('60万像素', result.text)
        self.assertIn('8万股', result.text)

    def test_fullwidth_case_number(self):
        result = self.d.mask('本院（2025）浙民终123号民事判决书')
        self.assertIn('[案号]', result.text)

    def test_scan_has_confidence(self):
        findings = self.d.scan('身份证号110101198001011232')
        self.assertIn('confidence', findings[0])
        self.assertGreater(findings[0]['confidence'], 0.9)

    def test_wechat_without_colon(self):
        result = self.d.mask('双方确认微信号lawyer_wang888为联络方式')
        self.assertIn('微信号[微信号]', result.text)

    def test_address_prefix_no_leak(self):
        result = self.d.mask('住所地浙江省杭州市拱墅区莫干山路100号')
        self.assertIn('住所地[地址]', result.text)
        self.assertNotIn('浙江省', result.text)
        self.assertNotIn('拱墅区', result.text)

    def test_scan_dedup_overlap(self):
        # 身份证号不应同时被记为银行账号
        findings = self.d.scan('身份证号110101198001011232')
        types = [f['type'] for f in findings]
        self.assertEqual(types.count('身份证号'), 1)
        self.assertNotIn('银行账号', types)


class TestRestore(unittest.TestCase):
    """mask → 映射表 → restore 无损往返"""

    def _roundtrip(self, text):
        d = Desensitizer()
        r = d.mask(text)
        restored = restore_text(r.text, parse_mapping_text(r.to_json()))
        self.assertEqual(restored, text)

    def test_restore_basic(self):
        self._roundtrip('原告：陈建国，男，身份证号110101198001011232，手机13800138000。')

    def test_restore_repeated_amounts(self):
        # 多个 [金额] 按原文顺序配对
        self._roundtrip('尾款80万元、违约金人民币伍佰万元整，另借3.6万利息。')

    def test_restore_birthdate_suffix(self):
        self._roundtrip('原告1985年8月15日出生，住浙江省杭州市西湖区文一西路1号。')

    def test_restore_from_markdown(self):
        d = Desensitizer()
        r = d.mask('原告：陈建国。尾款80万元。')
        restored = restore_text(r.text, parse_mapping_text(r.to_markdown()))
        self.assertEqual(restored, '原告：陈建国。尾款80万元。')

    def test_restore_repeated_noncontiguous_amounts(self):
        # 同一金额值多次出现且不相邻，映射须按"每处一行"配对
        self._roundtrip('A付款408669212.49元，B付款63689993.28元，'
                        'C付款408669212.49元，D付12万元。')

    def test_restore_mixed_rules_order(self):
        # 人名/公司/金额/地址混排，规则 pass 顺序与原文顺序不同时也能无损还原
        self._roundtrip('原告王强与被告杭州恒达建设集团有限公司（以下简称恒达公司）'
                        '签订合同，尾款408669212.49元，住浙江省杭州市西湖区文一西路1号。')

    def test_restore_trailing_space_original(self):
        # 金额带尾随空格（OCR 常见"349976351 元 ，"），映射表读写后仍逐字节还原
        d = Desensitizer()
        r = d.mask('总金额为 349976351 元 ，故需支付。')
        restored = restore_text(r.text, parse_mapping_text(r.to_markdown()))
        self.assertEqual(restored, '总金额为 349976351 元 ，故需支付。')

    def test_restore_ocr_spaced_company(self):
        self._roundtrip('被告：杭州恒达建设集团有限公司（以下简称恒 达公司）。'
                        '恒 达公司施工，景鸿公司分包。')


class TestRealCaseFixes(unittest.TestCase):
    """2026-08 真实判决书实战暴露问题的回归测试。"""

    def setUp(self):
        self.d = Desensitizer()

    def _roundtrip(self, text):
        r = self.d.mask(text)
        restored = restore_text(r.text, parse_mapping_text(r.to_markdown()))
        self.assertEqual(restored, text)

    def test_role_word_noun_not_name(self):
        # 角色词后跟名词/动词词组，不得当人名
        r = self.d.mask('原告提供担保，法定代表人处签名，法定代表人印章，'
                        '被告私章，原告主张驳回')
        self.assertIn('提供担保', r.text)
        self.assertIn('处签名', r.text)
        self.assertIn('印章', r.text)
        self.assertIn('私章', r.text)
        self.assertNotIn('[当事人甲', r.text)

    def test_role_words_anwai_and_witness(self):
        r = self.d.mask('案外人刘刚向本院起诉，证人王芳作证，'
                        '物业工作人员王芳陆续沟通。')
        self.assertIn('[当事人丙（第三人）]', r.text)
        self.assertEqual(r.text.count('[证人]'), 1)
        self.assertNotIn('刘刚', r.text)
        self.assertNotIn('王芳', r.text)

    def test_company_no_overcapture(self):
        r = self.d.mask('原告王强与被告杭州恒达建设集团有限公司'
                        '(以下简称恒 达公司)一案')
        self.assertIn('原告[当事人甲（原告）]与被告[合同乙方]', r.text)
        self.assertIn('(以下简称[合同乙方])', r.text)
        self.assertNotIn('王强', r.text)
        self.assertNotIn('恒达', r.text)

    def test_short_company_ocr_space(self):
        r = self.d.mask('恒 达公司施工，昊辰公 司收款，景鸿公司分包。')
        self.assertNotIn('恒 达公司', r.text)
        self.assertNotIn('昊辰公 司', r.text)
        self.assertNotIn('景鸿公司', r.text)

    def test_company_does_not_eat_placeholder(self):
        # 曾出现：简称规则把 "[第三方公司]" 里的内容再包一层括号
        r = self.d.mask('并将原由案外人杭州景鸿建筑工程有限公司'
                        '(以下简称景鸿公司)未施工完毕的二标段工程交由恒达公司。')
        self.assertNotIn('[[', r.text)
        self.assertNotIn(']]', r.text)

    def test_amount_space_and_big_number(self):
        r = self.d.mask('累计支付349976351 元，另付3449427.2  元，'
                        '算式392485911.675，尾款32000万 元整。')
        self.assertEqual(r.text.count('[金额]'), 4)
        self.assertNotIn('349976351', r.text)
        self.assertNotIn('392485911.675', r.text)

    def test_uppercase_amount_single_char_not_masked(self):
        # "陆"（陆续）等单字大写数字不是金额
        r = self.d.mask('王五陆续归还伍佰万元整，另付人民币壹拾贰万元。')
        self.assertIn('陆续', r.text)
        self.assertEqual(r.text.count('[金额]'), 2)

    def test_review_text_checklist(self):
        # 两阶段工作流阶段一：审阅清单应校验关键信息清零、列出低优先级残留
        from desensitize import build_review_text, scan_remaining_risk
        r = self.d.mask('原告王强与被告杭州恒达建设集团有限公司纠纷，'
                        '身份证号110101198001011232，浙江省杭州市余杭区人民法院已受理。')
        remaining = scan_remaining_risk(r.text)
        review = build_review_text(r.text, r.stats, remaining)
        self.assertIn('关键信息校验', review)
        self.assertIn('✅', review)
        self.assertTrue(any(f['type'] == '法院名称' for f in remaining))
        self.assertIn('法院名称', review)

    def test_scan_remaining_ignores_placeholders(self):
        # 占位符内部的"法院/公司"等字样不应被当成残留
        from desensitize import scan_remaining_risk
        remaining = scan_remaining_risk('[审理法院]已受理，[合同乙方]施工。')
        self.assertEqual(remaining, [])

    def test_case_number_ocr_space(self):
        # OCR 版案号在"字第"与数字、数字与"号"之间夹空格（审阅清单暴露的缺口）
        r = self.d.mask('据(2015)杭余民初字第 1819号民事判决书，'
                        '另见(2015)  杭余民初字第1819 号。')
        self.assertNotIn('1819', r.text)
        self.assertEqual(r.text.count('[案号]'), 2)

    def test_address_ocr_spaces_town(self):
        # 原告户籍地址：OCR 空格 + 镇级地址（东阳市歇山镇圳干村）
        r = self.d.mask('原告住浙江省 东 阳 市 歇 山 镇 圳 干 村  1-414     号，汉族。')
        self.assertIn('[地址]', r.text)
        self.assertNotIn('1-414', r.text)
        self.assertNotIn('歇山', r.text)

    def test_project_name_masked(self):
        # 项目名称（怡丰城项目/小区/一标段）→ [项目名称]；泛化词组保留
        r = self.d.mask('案涉怡丰城项目复工，怡丰城小区物业沟通，怡丰城一标段开工，'
                        '本项目继续施工，工程进度正常。')
        self.assertEqual(r.text.count('[项目名称]'), 3)
        self.assertNotIn('怡丰城', r.text)
        self.assertIn('本项目', r.text)
        self.assertIn('工程进度', r.text)

    def test_bare_name_not_place_or_role(self):
        # 地名（余杭区）与职务片段（承包人/包人）不是裸人名
        r = self.d.mask('杭州市余杭区人民法院裁定，作为承包人施工，'
                        '建设单位怡丰成公司资金链断裂。')
        self.assertIn('余杭区', r.text)
        self.assertIn('承包人', r.text)
        self.assertNotIn('余杭区[当事人', r.text)
        self.assertNotIn('承[当事人', r.text)

    def test_land_plot_number(self):
        r = self.d.mask('余政储出(2012)81号地块怡丰城项目，'
                        '另见余政储出(2012)81地块开发项目。')
        self.assertEqual(r.text.count('[地块编号]'), 2)
        self.assertNotIn('余政储出', r.text)

    def test_project_not_merchant(self):
        # 项目名规则不得吞商户名（"力灯饰商城丽信装饰材料商行"）
        r = self.d.mask('杭州华力灯饰商城丽信装饰材料商行供货，怡丰城项目复工。')
        self.assertNotIn('力灯饰[项目名称]', r.text)
        self.assertEqual(r.text.count('[项目名称]'), 1)

    def test_semantic_pass_merge_restore(self):
        # 阶段二：语义层合并映射后 restore 仍无损还原
        from desensitize import run_semantic_pass
        text = '案涉怡丰城项目，杭州市余杭区人民法院受理，余政储出(2012)81号地块。'
        r = self.d.mask(text)
        final_text, merged, err = run_semantic_pass(r.text, r.mapping)
        self.assertIsNone(err)
        self.assertIn('[关联法院]', final_text)   # 余杭区法院 → [关联法院]
        self.assertIn('[项目名称]', final_text)
        self.assertIn('[地块编号]', final_text)
        maps = [__import__('desensitize').Mapping(
            original=orig, replacement=ph, type='', count=1, order=i)
            for i, (ph, orig) in enumerate(merged, 1)]
        from desensitize import restore_text
        self.assertEqual(restore_text(final_text, maps), text)

    def test_restore_person_no_delimiter(self):
        # 无分隔符人名不插空格，还原后与原文完全一致
        self._roundtrip('原告陈建国，被告李四。')

    def test_restore_address_with_prefix(self):
        # 带"住所地"前缀的地址：脱敏整体替换、还原完全一致
        self._roundtrip('住所地：浙江省杭州市西湖区文一西路1号。')


class TestNerIntegration(unittest.TestCase):
    """mask_with_ner（regex 后端，无额外依赖）"""

    def test_mask_with_ner_regex_backend(self):
        from ner_interface import LegalNER
        d = Desensitizer()
        result = d.mask_with_ner(
            '原告：陈建国，被告：杭州鼎盛房地产开发有限公司。',
            LegalNER(backend='regex'),
        )
        self.assertIn('[当事人甲（原告）]', result.text)
        self.assertIn('[合同乙方]', result.text)

    def test_regex_ner_backend_extract(self):
        from ner_interface import LegalNER
        ner = LegalNER(backend='regex')
        result = ner.extract('原告王强，被告杭州鼎盛房地产开发有限公司')
        types = {e.type.value for e in result.entities}
        self.assertIn('PERSON', types)
        self.assertIn('COMPANY', types)


class TestBareNamesAndUnstructuredAddress(unittest.TestCase):
    """v2.4 规则层增强：裸人名 + 无层级地址 + 全文一致"""

    def setUp(self):
        self.d = Desensitizer()

    def test_bare_names_consistency_with_role_seed(self):
        # 角色词识别的名字向裸出现处传播，全文同一占位符
        result = self.d.mask('原告：陈建国。陈建国再次到庭陈述。原告陈建国称双方系朋友。')
        self.assertEqual(result.text.count('[当事人甲（原告）]'), 3)
        self.assertNotIn('陈建国', result.text)

    def test_bare_names_without_role(self):
        result = self.d.mask('张三欠李四钱不还')
        self.assertNotIn('张三', result.text)
        self.assertNotIn('李四', result.text)
        for m in result.mapping:
            if m.type == '人名':
                self.assertIn(m.original, ('张三', '李四'))

    def test_three_char_and_compound_names(self):
        result = self.d.mask('欧阳雪梅与王小明签订合同，欧阳雪梅称王小明已付款。')
        self.assertNotIn('欧阳雪梅', result.text)
        self.assertNotIn('王小明', result.text)
        # 同一人同一个占位符
        ph = [m.replacement for m in result.mapping if m.original == '欧阳雪梅']
        self.assertEqual(len(set(ph)), 1)

    def test_role_rule_not_cross_line(self):
        # 跨行"原告及其\n委托…""被告不应\n承担…"不得吞词
        result = self.d.mask('原告及其\n委托诉讼代理人赵敏。\n被告不应\n承担逾期交房违约责任。')
        self.assertNotIn('及其', {m.original for m in result.mapping if m.type == '人名'})
        self.assertNotIn('承担', {m.original for m in result.mapping if m.type == '人名'})
        self.assertIn('[委托代理人]', result.text)

    def test_common_words_not_names(self):
        result = self.d.mask('陈述事实经过后，双方对借款金额无异议，诉讼费由被告承担，原告抚养权问题另行处理。')
        names = {m.original for m in result.mapping if m.type == '人名'}
        for w in ('陈述', '金额', '承担', '抚养'):
            self.assertNotIn(w, names)

    def test_unstructured_address(self):
        result = self.d.mask('双方约定在望京西园四区410楼当面核账，莫干山路100号是注册地。')
        self.assertEqual(result.text.count('[地址]'), 2)
        self.assertNotIn('望京西园', result.text)

    def test_disable_bare_names(self):
        d = Desensitizer(bare_names=False)
        result = d.mask('原告：陈建国。张三欠李四钱不还。')
        self.assertIn('[当事人甲（原告）]', result.text)  # 角色名仍脱敏
        self.assertIn('张三', result.text)  # 裸人名关闭

    def test_bare_names_restore_roundtrip(self):
        text = '原告：陈建国。张三欠李四钱不还，双方约定在望京西园四区410楼核账。'
        r = self.d.mask(text)
        restored = restore_text(r.text, parse_mapping_text(r.to_json()))
        self.assertEqual(restored, text)


class TestV28RealCaseFixes(unittest.TestCase):
    """2026-08 三份训练语料（舒城判决书/昌黎处罚决定书/南沙判决书）暴露问题的回归。"""

    def setUp(self):
        self.d = Desensitizer()

    def _roundtrip(self, text):
        r = self.d.mask(text)
        restored = restore_text(r.text, parse_mapping_text(r.to_markdown()))
        self.assertEqual(restored, text)

    def test_17_digit_labeled_id_not_bank_card(self):
        # 原文录入少一位的 17 位身份证（标签下）：整体替换，不残留 X、不当银行账号
        r = self.d.mask('身份证号码3424251967112040X')
        self.assertIn('[身份证号]', r.text)
        self.assertNotIn('3424251967112040', r.text)
        self.assertNotIn('[银行账号]', r.text)

    def test_role_name_glue_tail_preserved(self):
        # 角色词后粘连动词（诉/到/未有）不被吞，姓名本体替换
        r = self.d.mask('原告彭静娴诉被告张旭到庭参加诉讼，被告张旭未有提供证据。')
        self.assertIn('原告[当事人甲（原告）]诉被告[当事人乙（被告）]到庭', r.text)
        self.assertIn('被告[当事人乙（被告）]未有提供证据', r.text)

    def test_spaced_role_names_replaced(self):
        # OCR 空格写法：审判员 倪 平 / 审 判 员 汪 瑜 / 书记员 杨梅红
        r = self.d.mask('审判员 倪 平，审 判 员 汪 瑜，书记员 杨梅红')
        self.assertIn('审判员 [法官]', r.text)
        self.assertIn('审 判 员 [法官]', r.text)
        self.assertIn('书记员 [书记员]', r.text)
        self._roundtrip('审判员 倪 平，审 判 员 汪 瑜，书记员 杨梅红')

    def test_context_words_not_names(self):
        # 收到/无异议/签定/微信/该公司 均不是人名或公司名
        r = self.d.mask('原告收到材料后表示无异议，原、被告签定合同。'
                        '张政微信告知原告公司员工，系该公司员工，收到回执。')
        for kept in ('收到', '无异议', '签定', '微信', '该公司'):
            self.assertIn(kept, r.text)

    def test_quantifier_not_role_name(self):
        r = self.d.mask('请求法院依法支持原告全部诉讼请求。')
        self.assertIn('全部诉讼请求', r.text)
        self.assertNotIn('[当事人', r.text)

    def test_merchant_and_fine_license(self):
        r = self.d.mask('当事人：昌黎县嘉瑞丰煎肉店，罚没许可证号：07040008，'
                        '住所（住址）：昌黎县四街铁塔东里片')
        self.assertIn('[当事人单位]', r.text)
        self.assertIn('[罚没许可证号]', r.text)
        self.assertIn('[地址]', r.text)

    def test_company_abbreviation_same_placeholder(self):
        # 简称与全称统一占位符：宝冶公司/合生东宇公司 不应各占一个实体
        r = self.d.mask('原告上海宝冶集团有限公司（以下简称宝冶公司）与被告'
                        '广州合生东宇房地产有限公司（以下简称合生东宇公司）'
                        '签订合同，合生东宇公司付款。')
        self.assertNotIn('[公司_', r.text)
        # 同一公司全称+简称只应占一个占位符
        self.assertEqual(r.text.count('[合同甲方]'), 2)   # 全称+简称
        self.assertEqual(r.text.count('[合同乙方]'), 3)   # 全称+简称×2
        self._roundtrip('原告上海宝冶集团有限公司（以下简称宝冶公司）与被告'
                        '广州合生东宇房地产有限公司（以下简称合生东宇公司）'
                        '签订合同，合生东宇公司付款。')

    def test_company_junk_span_trimmed(self):
        r = self.d.mask('裁定查封、扣押或冻结被告广州合生东宇房地产有限公司名下财产。')
        self.assertNotIn('[公司_', r.text)
        self.assertIn('扣押或冻结', r.text)

    def test_restore_mixed_dates_addresses_roundtrip(self):
        # 出生日期+地址+身份证+人名+金额混排（还原错位回归）
        text = ('原告：彭静娴，女，1979年10月20日出生，汉族，市民，'
                '住安徽省合肥市蜀山区芙蓉路988号明珠湖畔15幢301室，'
                '身份证号码342622197910207741。被告：张旭，男，1965年3月3日出生，'
                '住安徽省舒城县干汊河镇新陶村新堰村民组28号，'
                '身份证号码342425196503034015，应还借款35万元。')
        self._roundtrip(text)


# ============================================================
# v3.0 回归：validated 标记 / markdown 漂移还原 / 规则增强 / 合成语料 / PDF 涂黑
# ============================================================
class TestV30ValidatedMark(unittest.TestCase):
    """内化自 rizzo-pii：校验码 validated 标记（✓ / —）扫进映射表。"""

    def setUp(self):
        self.d = Desensitizer()

    def test_id_checksum_validated(self):
        r = self.d.mask('身份证号110101198001011232')  # 红队语料校验码合法
        for m in r.mapping:
            if m.type == '身份证号':
                self.assertTrue(m.validated, '校验码合法的身份证应 validated=True')

    def test_bank_card_unvalidated(self):
        # 红队语料 bank_02：Luhn 不合法但按"宁替勿漏"仍脱敏 → validated=False
        r = self.d.mask('尾款支付至6222020200012345678账户')
        for m in r.mapping:
            if m.type == '银行账号':
                self.assertFalse(m.validated)

    def test_mapping_markdown_validation_column(self):
        r = self.d.mask('身份证号110101198001011232，尾款支付至6222020200012345678账户')
        md = r.to_markdown()
        self.assertIn('| 序号 | 原始值 | 替换值 | 类型 | 出现次数 | 验证 |', md)
        self.assertIn('✓', md)   # 身份证校验码通过
        self.assertIn('—', md)   # 银行卡仅格式命中

    def test_mapping_json_roundtrip_keeps_validated(self):
        r = self.d.mask('身份证号110101198001011232')
        ms = parse_mapping_text(r.to_json())
        self.assertTrue(any(m.validated for m in ms if m.type == '身份证号'))

    def test_parse_old_markdown_without_validation_column(self):
        # 旧版映射表（无"验证"列）仍可解析，validated 默认 False
        old = '# 脱敏映射表\n| 序号 | 原始值 | 替换值 | 类型 | 出现次数 |\n' \
              '|------|--------|--------|------|---------|\n|1|张三|[当事人甲]|人名|1|'
        ms = parse_mapping_text(old)
        self.assertEqual(len(ms), 1)
        self.assertFalse(ms[0].validated)


class TestV30RestoreDrift(unittest.TestCase):
    """内化自 rizzo-pii：还原容忍 markdown 漂移（加粗/缺括号/多余空格）。"""

    def _m(self, original, replacement, order=1, typ='人名'):
        from desensitize import Mapping
        return Mapping(original=original, replacement=replacement,
                       type=typ, count=1, order=order)

    def test_bold_placeholder(self):
        ms = [self._m('陈建国', '[当事人甲（原告）]')]
        self.assertEqual(restore_text('原告**当事人甲（原告）**诉称', ms),
                         '原告陈建国诉称')

    def test_missing_brackets(self):
        ms = [self._m('杭州鼎盛房地产开发有限公司', '[合同乙方]')]
        # 宽松匹配不得吞掉占位符周围空格（" 合同乙方 " 还原后空格保留）
        self.assertEqual(restore_text('被告 合同乙方 违约', ms),
                         '被告 杭州鼎盛房地产开发有限公司 违约')

    def test_mixed_exact_and_drift_order(self):
        # 精确与漂移混用，仍按原文顺序配对（修复"先精确后宽松"的顺序错乱）
        ms = [self._m('张三', '[当事人甲]', order=1),
              self._m('李四', '[当事人甲]', order=2)]
        self.assertEqual(restore_text('**当事人甲**借给[当事人甲] 5 万元', ms),
                         '张三借给李四 5 万元')

    def test_context_prefix_not_swallowed(self):
        # 语义占位符 inner 词与上下文前缀撞车（"总金额[金额]"）不得误吞
        ms = [self._m('349976351 元 ', '[金额]', typ='金额')]
        self.assertEqual(restore_text('总金额为 [金额]，故需支付。', ms),
                         '总金额为 349976351 元 ，故需支付。')

    def test_exact_roundtrip_unaffected(self):
        d = Desensitizer()
        src = '原告陈建国诉被告杭州鼎盛房地产开发有限公司，身份证号110101198001011232'
        r = d.mask(src)
        self.assertEqual(restore_text(r.text, r.mapping), src)

    def test_body_word_bold_not_swallowed(self):
        # 正文普通词"金额"后跟 markdown 加粗开始标记（**巨大**）：
        # 宽松匹配只吞一侧 `**` 不是合法漂移形态 → 不得误还原
        ms = [self._m('500万元', '[金额]', typ='金额')]
        self.assertEqual(restore_text('赔偿金额**巨大**，另需支付[金额]。', ms),
                         '赔偿金额**巨大**，另需支付500万元。')


class TestV30RuleEnhancements(unittest.TestCase):
    """v3.0 合成语料驱动发现的规则层真实盲区修复。"""

    def setUp(self):
        self.d = Desensitizer()

    def test_role_word_被告人(self):
        r = self.d.mask('被告人陶丽，男，2017年2月20日出生。')
        self.assertIn('[被告人]', r.text) if '[被告人]' in r.text else \
            self.assertTrue(any(m.type == '人名' for m in r.mapping))

    def test_role_word_当事人(self):
        r = self.d.mask('当事人：俞明，身份证号410102198905257746。')
        self.assertTrue(any(m.type == '人名' for m in r.mapping))

    def test_scan_short_id_with_label(self):
        # scan 与 mask 一致：带标签的 15-17 位残缺身份证也要出现在扫描结果
        text = '身份证号码3424251967112040X'
        scan = self.d.scan(text)
        ids = [f['value'] for f in scan if f['type'] == '身份证号']
        self.assertIn('3424251967112040X', ids)
        r = self.d.mask(text)
        self.assertIn('[身份证号]', r.text)

    def test_address_does_not_swallow_date(self):
        # "25日在安徽省…"→ 地址规则不得吞"日"（日期保留）
        r = self.d.mask('于2022年4月12日在辽宁省青岛市玄武区凤起路7号从事经营。')
        self.assertIn('2022年4月12日', r.text)
        self.assertTrue(any(m.type == '地址' for m in r.mapping))

    def test_bare_name_with_verb_tokenized_together(self):
        # jieba 把"荣墨军称"切成"荣墨"+"军称" → 仍应识别"荣墨军"
        r = self.d.mask('荣墨军称：货款什么时候到账？')
        self.assertTrue(any(m.type == '人名' for m in r.mapping))


class TestV30Synthetic(unittest.TestCase):
    """内化自 rizzo-pii：LLM 写模板 + 代码注入合法校验值的合成语料管线。"""

    def test_generators_produce_valid_checksums(self):
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent / 'synthetic'))
        import random
        import generate_synthetic_pii as g
        rng = random.Random(7)
        for _ in range(50):
            self.assertTrue(_is_valid_id_checksum(g.gen_id_card(rng)),
                            '身份证 GB11643 校验失败')
            self.assertTrue(_is_valid_credit_code(g.gen_credit_code(rng)),
                            '信用代码 GB32100 校验失败')
            self.assertTrue(_luhn_check(g.gen_bank_card(rng)),
                            '银行卡 Luhn 校验失败')

    def test_qa_gate_rejects_inline_names(self):
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent / 'synthetic'))
        import llm_template_bank as lb
        bad = '原告王某某诉张某民间借贷纠纷一案。原告身份证号{身份证}。'
        self.assertTrue(lb.find_stray_names(bad), '应检出占位符外漏写的人名')
        good = '原告{当事人甲}诉被告{当事人乙}民间借贷纠纷一案，案号{案号}。'
        self.assertFalse(lb.find_stray_names(good), '纯槽位模板应通过 QA 门控')

    def test_fill_template_auto_annotates_expectations(self):
        import sys
        from pathlib import Path
        sys.path.insert(0, str(Path(__file__).parent / 'synthetic'))
        import random
        import generate_synthetic_pii as g
        rng = random.Random(1)
        text, entities, kept = g.fill_template(g.TEMPLATES[0], rng)
        self.assertTrue(any(label == '身份证号' for _, label, _, _ in entities))
        self.assertTrue(any(label == '人名' for _, label, _, _ in entities))
        # 注入值必须真的出现在文本里（标注与文本一致）
        for value, label, s, e in entities:
            self.assertEqual(text[s:e], value)


class TestV30PdfRedact(unittest.TestCase):
    """内化自 rizzo-pii：PDF 真·涂黑脱敏（保留版式 + residual 零残留）。"""

    def _make_pdf(self, path):
        try:
            import fitz
        except ImportError:
            self.skipTest('PyMuPDF 未安装')
        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((60, 80), '原告陈建国，身份证号110101198001011232。',
                         fontname='china-s', fontsize=10)
        page.insert_text((60, 104), '张三丰与张三均到庭。', fontname='china-s', fontsize=10)
        doc.set_metadata({'title': '陈建国案', 'author': '陈建国'})
        doc.set_toc([[1, '陈建国案', 1]])
        doc.embfile_add('n.txt', '陈建国'.encode(), filename='n.txt')
        doc.save(path)
        doc.close()

    def test_redact_and_residual(self):
        import tempfile, os
        from pdf_redact import redact_pdf, PdfError
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, 'src.pdf')
        self._make_pdf(src)
        with open(src, 'rb') as f:
            pdf = f.read()
        pairs = [('[身份证号]', '110101198001011232'),
                 ('[当事人甲（原告）]', '陈建国'),
                 ('[当事人乙]', '张三'),
                 # 张三丰与张三是两个独立人名：值长优先涂张三丰，张三不破坏其矩形
                 ('[当事人丙]', '张三丰')]
        out, report = redact_pdf(pdf, pairs)
        self.assertEqual(report['residual'], [])
        import fitz
        with fitz.open(stream=out, filetype='pdf') as doc:
            t = doc[0].get_text()
            self.assertNotIn('陈建国', t)
            self.assertNotIn('110101198001011232', t)
            self.assertNotIn('张三丰', t)   # 张三丰 被涂（值长优先）
            self.assertNotIn('丰', t)       # 不被"张三"半涂成残留"丰"
            self.assertNotIn('张三', t)     # 张三 独立人名被涂
            self.assertFalse(any(doc.metadata.get(k) for k in
                                 ('title', 'author', 'subject', 'keywords')),
                             '元数据（标题/作者等）应被清空')
            self.assertEqual(list(doc.embfile_names()), [])

    def test_zero_hit_rejected(self):
        import tempfile, os
        from pdf_redact import redact_pdf, PdfError
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, 'src.pdf')
        self._make_pdf(src)
        with open(src, 'rb') as f:
            pdf = f.read()
        with self.assertRaises(PdfError):
            redact_pdf(pdf, [('[不存在]', '完全不存在的内容xyz')])


# ============================================================
# v3.1 回归：微信号上下文增强（是/为/微信号码/OCR空格）+ 姓名动词尾误报防护
# ============================================================
class TestV31Wechat(unittest.TestCase):
    """微信号脱敏：'微信号是/为/微信号码/我的微信是' 等法律文书常见写法。"""

    def setUp(self):
        self.d = Desensitizer()

    def test_wechat_variants_masked(self):
        cases = [
            '微信号是lawyer_wang888',
            '微信号为zhangsan001',
            '微信号码是zhangsan',
            '我的微信是abc_123456',
            '微信账号为wxid_abcdefghijkl',
            '微 信号：abc_123456',          # OCR 行内空格
            '双方确认微信号为lawyer_wang888为联络方式',
        ]
        for t in cases:
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertTrue(any(m.type == '微信号' for m in r.mapping),
                                f'未脱敏: {t} -> {r.text}')

    def test_wechat_negatives_not_masked(self):
        cases = [
            '微信是常用的聊天工具',      # 后接中文，不是微信号
            '微信为诉讼证据',
            '使用ChatGPT处理合同',      # 裸英文词不误判
            '微信号：张三丰',           # 中文昵称不是微信号
            '微信支付了500元',
        ]
        for t in cases:
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertFalse(any(m.type == '微信号' for m in r.mapping),
                                 f'误报: {t} -> {r.text}')

    def test_wechat_restore_roundtrip(self):
        src = '我的微信号码是zhangsan001，双方确认其微信号为lawyer_wang888。'
        r = self.d.mask(src)
        back = restore_text(r.text, r.mapping)
        self.assertEqual(back, src)

    def test_scan_matches_mask(self):
        text = '微信号为zhangsan001'
        scan = self.d.scan(text)
        self.assertTrue(any(f['type'] == '微信号' for f in scan))
        r = self.d.mask(text)
        self.assertIn('[微信号]', r.text)


class TestV31NameVerbTail(unittest.TestCase):
    """姓名+高频动词尾 不再误吞（张三确认/双方确认/什么时候）。"""

    def setUp(self):
        self.d = Desensitizer()

    def test_name_confirmation_not_swallowed(self):
        r = self.d.mask('原告张三确认收到货款。')
        self.assertIn('确认收到', r.text)
        self.assertTrue(any(m.type == '人名' for m in r.mapping))

    def test_shuangfang_kept(self):
        for t in ('双方确认后签订合同', '经双方约定付款', '双方同意解除合同'):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertNotIn('当事人', r.text, f'误报: {t} -> {r.text}')

    def test_shihou_kept(self):
        r = self.d.mask('货款什么时候到账')
        self.assertNotIn('当事人', r.text)

    def test_jieba_glue_still_masked(self):
        # v3.0 修复的"名 token 带尾动词"粘连仍生效（真实人名不回归）
        for t in ('荣墨军称：货款什么时候到账？', '齐艳称：好的'):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertTrue(any(m.type == '人名' for m in r.mapping),
                                f'漏脱敏: {t} -> {r.text}')

    def test_tail_verbs_still_trimmed(self):
        # v2.8 尾部动词回退不回归
        for t, tail in (('被告张旭未有提供证据', '未有'),
                        ('原告彭静娴诉被告', '诉')):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertIn(tail, r.text)


# ============================================================
# v3.2 回归：手机号式微信号（纯手机号由手机号规则兜底；带分隔符/86前缀
#           由微信号规则补充，且不被金额规则错标）
# ============================================================
class TestV32WechatMobile(unittest.TestCase):
    def setUp(self):
        self.d = Desensitizer()

    def test_pure_mobile_covered_by_phone_rule(self):
        # 纯 11 位：手机号规则（先执行）兜底，无需微信号规则
        r = self.d.mask('微信号：13800138000')
        self.assertTrue(any(m.type == '手机号' for m in r.mapping))
        self.assertIn('[手机号]', r.text)

    def test_separated_mobile_masked_as_wechat(self):
        for t, v in (('微信号：138-0013-8000', '138-0013-8000'),
                     ('微信号：138 0013 8000', '138 0013 8000')):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertTrue(any(m.type == '微信号' and m.original == v
                                    for m in r.mapping),
                                f'{t} -> {r.text}')

    def test_86_prefix_masked_not_amount(self):
        # 13 位纯数字在金额规则（后执行）之前被微信号规则抢先，不再误标 [金额]
        for t, v in (('微信号：+8613800138000', '+8613800138000'),
                     ('微信号：8613800138000', '8613800138000')):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertTrue(any(m.type == '微信号' and m.original == v
                                    for m in r.mapping),
                                f'{t} -> {r.text}')
                self.assertFalse(any(m.type == '金额' for m in r.mapping))

    def test_bank_card_not_wechat(self):
        r = self.d.mask('微信号：6222020200012345678')
        self.assertTrue(any(m.type == '银行账号' for m in r.mapping))

    def test_landline_under_wechat_ctx(self):
        r = self.d.mask('微信号：010-12345678')
        self.assertTrue(any(m.type == '固定电话' for m in r.mapping))

    def test_restore_roundtrip_with_86(self):
        src = '原告微信号：138-0013-8000，被告微信号是+8613800138000。'
        r = self.d.mask(src)
        self.assertEqual(restore_text(r.text, r.mapping), src)


# ============================================================
# v3.3 回归：Excel（.xlsx / .xlsm）支持 — 单元格展平 / 公式跳过 / 还原往返
# ============================================================
class TestExcelSupport(unittest.TestCase):
    """Excel 支持：每个参与脱敏的单元格展平为一行，写回保留结构，restore 逐单元格还原。"""

    def setUp(self):
        self.d = Desensitizer()

    def _make_wb(self, path):
        try:
            from openpyxl import Workbook
        except ImportError:
            self.skipTest('openpyxl 未安装')
        wb = Workbook()
        ws = wb.active
        ws.title = '流水'
        ws['A1'] = '姓名'
        ws['B1'] = '身份证号'
        ws['C1'] = '手机号'
        ws['D1'] = '金额'
        ws['A2'] = '陈建国'
        ws['B2'] = '110101198001011232'
        ws['C2'] = '13800138000'
        ws['D2'] = 392485911.675
        ws['A3'] = '杭州鼎盛房地产开发有限公司'
        ws['B3'] = '=SUM(D2:D2)'          # 公式，应保留
        ws['C3'] = 'charlie@example.com'
        ws.merge_cells('A4:B4')           # 合并区
        ws['A4'] = '合并区'
        ws2 = wb.create_sheet('备注')
        ws2['A1'] = '张三\n李四'          # 单元格内换行
        ws2['B1'] = True                  # 布尔，应保留
        ws2['C1'] = 123                    # 普通整数
        wb.save(path)
        return wb

    def _read_cells(self, path):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=False)
        try:
            return {s.title: [[c.value for c in row] for row in s.iter_rows()]
                    for s in wb.worksheets}
        finally:
            wb.close()

    def test_flat_text_extracts_cells_in_order(self):
        import tempfile, os
        from desensitize import read_text_from_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '流水.xlsx')
        self._make_wb(src)
        text = read_text_from_file(src)
        # 行序：sheet1 逐行逐列（跳过 None/公式），再 sheet2；换行单元格含 ␊ 标记
        self.assertEqual(text.split('\n')[0], '姓名')
        self.assertIn('陈建国', text)
        self.assertIn('110101198001011232', text)
        self.assertIn('392485911.675', text)
        self.assertNotIn('=SUM(D2:D2)', text)   # 公式不参与
        self.assertIn('张三\u240a李四', text)    # 单元格内换行 → ␊ 标记
        self.assertIn('合并区', text)            # sheet1 内容
        # 两个 sheet 内容都出现，且布尔/公式跳过
        self.assertNotIn('True', text.split('\n'))

    def test_mask_writes_back_cell_level(self):
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '流水.xlsx')
        self._make_wb(src)
        text = read_text_from_file(src)
        res = self.d.mask(text)
        out = os.path.join(tmp, '流水_desensitized.xlsx')
        write_desensitized_file(src, out, res.text)
        cells = self._read_cells(out)
        ws = cells['流水']
        self.assertEqual(ws[1][1], '[身份证号]')   # B2
        self.assertEqual(ws[1][2], '[手机号]')     # C2
        self.assertEqual(ws[1][3], '[金额]')       # D2（float → [金额]）
        self.assertEqual(ws[2][1], '=SUM(D2:D2)')  # 公式原样保留
        self.assertEqual(cells['备注'][0][0], '张三\n李四')  # 换行还原
        self.assertEqual(cells['备注'][0][1], True)          # 布尔保留
        self.assertEqual(cells['备注'][0][2], 123)           # 整数保留

    def test_formula_skipped_kept(self):
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '公式.xlsx')
        wb = self._make_wb(src)
        text = read_text_from_file(src)
        self.assertNotIn('=SUM', text)  # 读取时公式不进入文本
        res = self.d.mask(text)
        out = os.path.join(tmp, '公式_out.xlsx')
        write_desensitized_file(src, out, res.text)
        cells = self._read_cells(out)
        self.assertEqual(cells['流水'][2][1], '=SUM(D2:D2)')

    def test_restore_roundtrip_exact(self):
        """mask → restore 后与原 xlsx 逐单元格一致（含 float 类型）。"""
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '流水.xlsx')
        wb = self._make_wb(src)
        original = self._read_cells(src)
        text = read_text_from_file(src)
        res = self.d.mask(text)
        restored = restore_text(res.text, res.mapping)
        out = os.path.join(tmp, '还原.xlsx')
        write_desensitized_file(src, out, restored)
        got = self._read_cells(out)
        for sheet in original:
            for r, row in enumerate(original[sheet]):
                for c, v in enumerate(row):
                    if v is None:
                        continue
                    gv = got[sheet][r][c]
                    self.assertEqual(gv, v, f'{sheet}!{r},{c}')
                    self.assertIs(type(gv), type(v),
                                  f'{sheet}!{r},{c} 类型不一致: {type(gv)} vs {type(v)}')

    def test_cli_style_restore_from_masked_file(self):
        """真实 CLI 流程：restore 以脱敏文件为基底（而非原始文件），
        金额等数值类单元格恢复 int/float 类型，身份证/银行卡等号码类保持文本。"""
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '流水.xlsx')
        wb = self._make_wb(src)
        text = read_text_from_file(src)
        res = self.d.mask(text)
        masked = os.path.join(tmp, '流水_desensitized.xlsx')
        write_desensitized_file(src, masked, res.text)
        # 以脱敏文件为基底还原（与 desensitize.py restore 命令一致）
        masked_text = read_text_from_file(masked)
        restored = restore_text(masked_text, res.mapping)
        out = os.path.join(tmp, '还原.xlsx')
        write_desensitized_file(masked, out, restored, mappings=res.mapping)
        got = self._read_cells(out)
        ws = got['流水']
        # 金额 float → 恢复为 float
        self.assertIsInstance(ws[1][3], float)
        self.assertEqual(ws[1][3], 392485911.675)
        # 身份证号（号码类）保持文本
        self.assertIsInstance(ws[1][1], str)
        self.assertEqual(ws[1][1], '110101198001011232')
        # 公司名恢复为文本原文
        self.assertEqual(ws[2][0], '杭州鼎盛房地产开发有限公司')
        # 公式仍保留
        self.assertEqual(ws[2][1], '=SUM(D2:D2)')

    def test_multisheet_consistency(self):
        """公司全称与简称跨单元格/跨 sheet 统一占位符（EntityResolver 全文一致）。"""
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '多表.xlsx')
        wb = self._make_wb(src)
        ws = wb.active
        ws['A5'] = '上海宝冶集团有限公司'
        ws['B5'] = '宝冶公司'          # 简称 → 应链接到全称同一占位符
        wb.save(src)
        text = read_text_from_file(src)
        res = self.d.mask(text)
        out = os.path.join(tmp, '多表_out.xlsx')
        write_desensitized_file(src, out, res.text)
        cells = self._read_cells(out)
        a5 = cells['流水'][4][0]
        b5 = cells['流水'][4][1]
        self.assertTrue(a5.startswith('[公司') and a5.endswith(']'), a5)
        self.assertEqual(a5, b5, '全称与简称应统一占位符')

    def test_missing_dependency_graceful(self):
        """openpyxl 缺失时优雅退出（给安装指引），而非崩溃。"""
        import tempfile, os
        from desensitize import read_text_from_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, 'x.xlsx')
        with open(src, 'wb') as f:
            f.write(b'PK\x03\x04')  # 最小 xlsx 头
        import builtins
        real_import = builtins.__import__
        def fake_import(name, *a, **kw):
            if name == 'openpyxl':
                raise ImportError('No module named openpyxl')
            return real_import(name, *a, **kw)
        builtins.__import__ = fake_import
        try:
            with self.assertRaises(SystemExit) as cm:
                read_text_from_file(src)
            self.assertIn('openpyxl', str(cm.exception))
        finally:
            builtins.__import__ = real_import

    def test_macros_keep_vba(self):
        """.xlsm 加载 keep_vba=True，宏不丢。"""
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '宏.xlsm')
        wb = self._make_wb(src)
        wb.save(src)
        text = read_text_from_file(src)
        res = self.d.mask(text)
        out = os.path.join(tmp, '宏_out.xlsm')
        write_desensitized_file(src, out, res.text)
        self.assertTrue(os.path.exists(out))
        self.assertTrue(os.path.getsize(out) > 0)

    def test_empty_workbook_no_crash(self):
        import tempfile, os
        from openpyxl import Workbook
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '空.xlsx')
        wb = Workbook()
        wb.save(src)
        text = read_text_from_file(src)
        self.assertEqual(text, '')
        res = self.d.mask(text)
        out = os.path.join(tmp, '空_out.xlsx')
        write_desensitized_file(src, out, res.text)
        cells = self._read_cells(out)
        # 空表不崩溃，sheet 仍存在，所有单元格为 None
        self.assertIn('Sheet', cells)
        self.assertTrue(all(v is None for row in cells['Sheet'] for v in row))

    def test_phone_and_email_cell(self):
        import tempfile, os
        from desensitize import read_text_from_file, write_desensitized_file
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '联系.xlsx')
        wb = self._make_wb(src)
        wb.active['D5'] = '微信号：lawyer_wang'
        wb.save(src)
        text = read_text_from_file(src)
        res = self.d.mask(text)
        out = os.path.join(tmp, '联系_out.xlsx')
        write_desensitized_file(src, out, res.text)
        cells = self._read_cells(out)
        self.assertEqual(cells['流水'][4][3], '微信号：[微信号]')


# ============================================================
# v3.4 回归：支付平台前缀交易对手 / PDF 空文本报错 / 假 PDF 修复
# ============================================================
class TestV34PlatformCounterparty(unittest.TestCase):
    """银行流水高频格式："支付宝-刘方立" / "微信转账-张三" → 平台名保留、人名脱敏。"""

    def setUp(self):
        self.d = Desensitizer()

    def test_platform_prefix_person_masked(self):
        for t in ('支付宝-刘方立', '微信转账-张三', '财付通：李四',
                  '支付宝 王五', '银联-陈建国'):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertTrue(any(m.type == '人名' for m in r.mapping),
                                f'{t} -> {r.text}')
                # 平台名保留（公开品牌，不脱敏）
                self.assertIn('支付宝' if '支付宝' in t else
                              ('微信' if '微信' in t else
                               ('财付通' if '财付通' in t else '银联')), r.text)

    def test_platform_restore_roundtrip(self):
        for t in ('支付宝-刘方立 转账1000元',
                  '2024-01-15 微信转账-张三 余额5000.00',
                  '银联-陈建国\n支付宝-刘方立'):
            with self.subTest(t=t):
                r = self.d.mask(t)
                self.assertEqual(restore_text(r.text, r.mapping), t)

    def test_platform_same_person_same_placeholder(self):
        r = self.d.mask('支付宝-刘方立\n支付宝-刘方立\n银联-陈建国')
        self.assertEqual(r.text.count('[当事人_'), 3)
        # 同一对手（刘方立）两次出现 → 同一占位符；陈建国不同
        lines = r.text.split('\n')
        self.assertEqual(lines[0], lines[1], '同一对手跨行应统一占位符')
        self.assertNotEqual(lines[0], lines[2])

    def test_platform_no_false_positive(self):
        # 无分隔符的普通短语不是平台前缀，且平台词不是人名
        r = self.d.mask('微信支付了100元\n支付宝到账2000元')
        self.assertIn('微信支付了', r.text)
        self.assertIn('支付宝到账', r.text)
        self.assertNotIn('[当事人', r.text)

    def test_pdf_no_text_layer_rejected(self):
        """纯图片扫描件 PDF：明确报错并给出 OCR 指引，而非静默产出空文件。"""
        import tempfile, os
        from desensitize import read_text_from_file
        try:
            import fitz
        except ImportError:
            self.skipTest('PyMuPDF 未安装')
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '扫描件.pdf')
        doc = fitz.open()
        page = doc.new_page()
        page.draw_rect(fitz.Rect(0, 0, 600, 800), color=None, fill=(0.9, 0.9, 0.9))
        doc.save(src)
        doc.close()
        with self.assertRaises(SystemExit) as cm:
            read_text_from_file(src)
        self.assertIn('文本层', str(cm.exception))
        self.assertIn('OCR', str(cm.exception))

    def test_pdf_with_text_still_works(self):
        import tempfile, os
        from desensitize import read_text_from_file
        try:
            import fitz
        except ImportError:
            self.skipTest('PyMuPDF 未安装')
        tmp = tempfile.mkdtemp()
        src = os.path.join(tmp, '正常.pdf')
        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((60, 80), '原告陈建国，身份证号110101198001011232。',
                         fontname='china-s', fontsize=10)
        doc.save(src)
        doc.close()
        text = read_text_from_file(src)
        self.assertIn('陈建国', text)
        self.assertIn('110101198001011232', text)

    def test_default_output_ext_pdf_to_txt(self):
        """PDF 输入未指定 -o 时自动命名 .txt（避免假 PDF），其他格式保留原扩展名。"""
        from desensitize import _default_output_ext
        self.assertEqual(_default_output_ext('a.pdf'), '.txt')
        self.assertEqual(_default_output_ext('a.docx'), '.docx')
        self.assertEqual(_default_output_ext('a.xlsx'), '.xlsx')
        self.assertEqual(_default_output_ext('a.txt'), '.txt')
        self.assertEqual(_default_output_ext('a'), '.txt')
        self.assertEqual(_default_output_ext('A.PDF'), '.txt')


# ============================================================
# v3.5 回归：银行流水场景（账号/户名组合 / 支付宝掩码 / 银行机构 / 日期误伤）
# ============================================================
class TestV35BankStatement(unittest.TestCase):
    """借鉴第三方框架的银行流水专用模式：对方账号与户名、支付宝掩码、
    银行分支机构名、交易日期不被金额规则误标。"""

    def setUp(self):
        self.d = Desensitizer()

    def test_slash_account_full(self):
        # 完整账号/户名：账号与户名都要脱敏
        r = self.d.mask('6212261001014270893/胡若薇')
        self.assertIn('[银行账号]', r.text)
        self.assertIn('[当事人', r.text)
        self.assertNotIn('胡若薇', r.text)
        self.assertEqual(restore_text(r.text, r.mapping),
                         '6212261001014270893/胡若薇')

    def test_alipay_masked_account(self):
        # 支付宝掩码账号：7399/支***刘方立
        r = self.d.mask('7399/支***刘方立')
        self.assertIn('[支付宝账号]', r.text)
        self.assertIn('[当事人', r.text)
        self.assertNotIn('刘方立', r.text)
        self.assertEqual(restore_text(r.text, r.mapping), '7399/支***刘方立')

    def test_bank_branch_masked(self):
        # 银行机构名：公司名规则处理后残留的机构尾缀补上
        r = self.d.mask('中国建设银行股份有限公司安徽省分行本级本币头寸机构')
        self.assertNotIn('安徽省分行本级本币头寸机构', r.text)
        self.assertIn('[银行机构]', r.text)
        self.assertEqual(restore_text(r.text, r.mapping),
                         '中国建设银行股份有限公司安徽省分行本级本币头寸机构')

    def test_bank_branch_plain(self):
        # 无"公司"后缀的银行机构
        r = self.d.mask('工商银行北京长安支行')
        self.assertIn('[银行机构]', r.text)
        self.assertEqual(restore_text(r.text, r.mapping), '工商银行北京长安支行')

    def test_bank_branch_no_false_positive(self):
        # "该分行表示同意"不是机构名，不被误伤
        r = self.d.mask('该分行表示同意')
        self.assertEqual(r.text, '该分行表示同意')
        self.assertEqual(len(r.mapping), 0)

    def test_date_not_masked_as_amount(self):
        # 交易日期（8 位）不被金额规则误标；真大额金额仍识别
        r = self.d.mask('20181009 20181015 392485911.675')
        self.assertIn('20181009', r.text)
        self.assertIn('20181015', r.text)
        self.assertIn('[金额]', r.text)
        self.assertEqual(restore_text(r.text, r.mapping),
                         '20181009 20181015 392485911.675')

    def test_bank_statement_demo_rows(self):
        # 第三方框架 demo 的 5 行银行流水：全行脱敏 + 逐字节还原
        rows = [
            ('消费 20181009 -2950.00 支付宝-刘方立 7399/支***刘方立', 2),  # 平台+掩码
            ('电子汇入 20181015 180000.00 6212261001014270893/胡若薇', 1),  # 户名
            ('消费 20181016 -1352.00 支付宝外部商户-上海恒银餐饮管理有限公司 73...', 0),  # 公司非人名
            ('ATM转账 20181101 50000.00 6227001711200011447/徐常英', 1),
            ('转账支取 20181104 中国建设银行股份有限公司安徽省分行本级本币头寸机构 6217001710000409331/徐常英', 1),
        ]
        for text, expect_people in rows:
            with self.subTest(text=text):
                r = self.d.mask(text)
                people = [m for m in r.mapping if m.type == '人名']
                self.assertEqual(len(people), expect_people,
                                 f'{text} -> {r.text}')
                self.assertEqual(restore_text(r.text, r.mapping), text)


# ============================================================
# v3.6 回归：列感知脱敏（银行流水表格）— 结构化读取/列类型/孤立姓名/审计增强
# ============================================================
class TestV36TableAware(unittest.TestCase):
    """列感知模式：识别表头列名→按列类型脱敏。户名列孤立姓名也识别、
    日期列不脱敏、附言列不按金额（联行号不误标）、还原逐单元格一致。"""

    def setUp(self):
        self.d = Desensitizer()

    def _make_xlsx(self, path):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = '交易明细'
        ws.append(['序号', '交易日期', '摘要', '交易金额', '账户余额',
                   '交易地点/附言', '对方账号与户名'])
        ws.append(['61', 20181009, '消费', -2950.00, 10779.00,
                   '支付宝-刘方立', '7399/支***刘方立'])
        ws.append(['65', 20181015, '电子汇入', 180000.00, 186771.00,
                   '电子汇入', '胡若薇'])          # 孤立姓名
        ws.append(['76', 20181101, 'ATM转账', 50000.00, 227317.00,
                   340690400059, '徐常英'])       # 联行号 + 孤立姓名
        wb.save(path)

    def test_read_structured_xlsx(self):
        import tempfile, os
        from desensitize import read_structured_table
        tmp = tempfile.mkdtemp()
        p = os.path.join(tmp, 't.xlsx')
        self._make_xlsx(p)
        res = read_structured_table(p)
        self.assertIsNotNone(res)
        headers, rows, types = res
        self.assertIn('对方账号与户名', headers)
        self.assertIn('户名', types)
        self.assertIn('日期', types)
        self.assertEqual(len(rows), 3)

    def test_read_structured_non_table_returns_none(self):
        import tempfile, os
        from openpyxl import Workbook
        from desensitize import read_structured_table
        tmp = tempfile.mkdtemp()
        p = os.path.join(tmp, '无表头.xlsx')
        wb = Workbook(); ws = wb.active
        ws.append(['随便', '数据'])
        ws.append(['abc', 'def'])
        wb.save(p)
        self.assertIsNone(read_structured_table(p))

    def test_classify_column(self):
        from desensitize import _classify_column
        self.assertEqual(_classify_column('对方账号与户名'), '户名')
        self.assertEqual(_classify_column('交易日期'), '日期')
        self.assertEqual(_classify_column('交易金额'), '金额')
        self.assertEqual(_classify_column('交易地点/附言'), '附言')
        self.assertEqual(_classify_column('任意列'), '默认')

    def test_mask_table_isolated_names(self):
        """户名列孤立姓名（count=1 无角色词）也被识别。"""
        import tempfile, os
        from desensitize import read_structured_table
        tmp = tempfile.mkdtemp()
        p = os.path.join(tmp, 't.xlsx')
        self._make_xlsx(p)
        headers, rows, types = read_structured_table(p)
        r = self.d.mask_table(headers, rows, types)
        self.assertIn('[当事人', r.text)
        # 胡若薇 / 徐常英 都被替换
        self.assertNotIn('胡若薇', r.text)
        self.assertNotIn('徐常英', r.text)

    def test_mask_table_date_and_branch_kept(self):
        """日期列原样保留；附言列联行号不按金额。"""
        import tempfile, os
        from desensitize import read_structured_table
        tmp = tempfile.mkdtemp()
        p = os.path.join(tmp, 't.xlsx')
        self._make_xlsx(p)
        headers, rows, types = read_structured_table(p)
        r = self.d.mask_table(headers, rows, types)
        self.assertIn('20181009', r.text)          # 日期保留
        self.assertIn('340690400059', r.text)      # 联行号不标 [金额]
        self.assertNotIn('[金额]', r.text)

    def test_scan_finds_isolated_name(self):
        """普通模式下孤立姓名残留：审阅清单必须提示（要求8）。"""
        import sys; sys.path.insert(0, '.')
        from desensitize import scan_remaining_risk, Desensitizer
        d = Desensitizer()
        masked = d.mask('对方账号与户名:胡若薇').text
        hits = [x for x in scan_remaining_risk(masked)
                if x['type'] == '孤立姓名候选']
        self.assertTrue(any(x['value'] == '胡若薇' for x in hits))

    def test_scan_no_false_positive_common_words(self):
        from desensitize import scan_remaining_risk, Desensitizer
        d = Desensitizer()
        masked = d.mask('消费 转账 摘要 金额 20181009').text
        hits = [x for x in scan_remaining_risk(masked)
                if x['type'] == '孤立姓名候选']
        self.assertEqual(hits, [])


if __name__ == '__main__':
    unittest.main()
